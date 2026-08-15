"""把财报数据与计算出的指标导出为 Excel（.xlsx）文件。

生成 4 个工作表：
1. 核心指标 —— 公司基础信息 + 最新一期核心指标 + 近 N 年趋势
2. 利润表
3. 资产负债表
4. 现金流量表
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# 东财英文字段名 -> 中文表头（金额字段，导出时统一换算为「亿元」）
INCOME_CN = {
    "REPORT_DATE": "报告期",
    "TOTAL_OPERATE_INCOME": "营业总收入(亿元)",
    "OPERATE_INCOME": "营业收入(亿元)",
    "OPERATE_COST": "营业成本(亿元)",
    "OPERATE_PROFIT": "营业利润(亿元)",
    "TOTAL_PROFIT": "利润总额(亿元)",
    "NETPROFIT": "净利润(亿元)",
    "PARENT_NETPROFIT": "归母净利润(亿元)",
    "DEDUCT_PARENT_NETPROFIT": "扣非归母净利润(亿元)",
}

BALANCE_CN = {
    "REPORT_DATE": "报告期",
    "TOTAL_ASSETS": "资产总计(亿元)",
    "TOTAL_LIABILITIES": "负债合计(亿元)",
    "TOTAL_EQUITY": "股东权益合计(亿元)",
    "TOTAL_PARENT_EQUITY": "归母股东权益(亿元)",
    "MONETARYFUNDS": "货币资金(亿元)",
    "GOODWILL": "商誉(亿元)",
    "TOTAL_CURRENT_ASSETS": "流动资产合计(亿元)",
    "TOTAL_CURRENT_LIAB": "流动负债合计(亿元)",
    "INVENTORY": "存货(亿元)",
    "ACCOUNTS_RECE": "应收账款(亿元)",
}

CASHFLOW_CN = {
    "REPORT_DATE": "报告期",
    "NETCASH_OPERATE": "经营现金流净额(亿元)",
    "NETCASH_INVEST": "投资现金流净额(亿元)",
    "NETCASH_FINANCE": "筹资现金流净额(亿元)",
}

_AMOUNT_FIELDS = set(INCOME_CN) | set(BALANCE_CN) | set(CASHFLOW_CN) - {"REPORT_DATE"}

_HEADER_FILL = PatternFill("solid", fgColor="2F6FED")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)


def _to_num(v):
    """把单元格值转成数字（无法解析则返回原值）。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    import re
    s = str(v).replace(",", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else v


def _style_header(ws, row_idx, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_overview(ws, data, indicators):
    """Sheet 1：核心指标。"""
    ci = data.get("company_info") or {}
    latest = indicators.get("latest") or {}
    history = indicators.get("history") or []

    # 标题与公司信息
    ws["A1"] = "%s（%s）核心财务指标" % (data.get("name") or "公司", data.get("security_code", ""))
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "所属行业：%s    主营业务：%s" % (ci.get("industry") or "—", ci.get("main_business") or "—")
    ws["A3"] = "数据抓取时间：%s    数据来源：%s" % (data.get("fetched_at") or "—", data.get("source") or "—")

    def pct(v):
        return None if v is None else round(v * 100, 2)

    # 最新一期核心指标（键值对）
    kv = [
        ("报告期", latest.get("report_date")),
        ("营业收入(亿元)", round(latest["revenue"] / 1e8, 2) if latest.get("revenue") is not None else None),
        ("营业收入同比(%)", pct(latest.get("revenue_yoy"))),
        ("归母净利润(亿元)", round(latest["parent_net_profit"] / 1e8, 2) if latest.get("parent_net_profit") is not None else None),
        ("归母净利润同比(%)", pct(latest.get("parent_net_profit_yoy"))),
        ("毛利率(%)", pct(latest.get("gross_margin"))),
        ("净利率(%)", pct(latest.get("net_margin"))),
        ("ROE 归母(%)", pct(latest.get("roe"))),
        ("总资产周转率(次)", latest.get("asset_turnover")),
        ("权益乘数", latest.get("equity_multiplier")),
        ("资产负债率(%)", pct(latest.get("debt_ratio"))),
        ("流动比率", latest.get("current_ratio")),
        ("经营现金流净额(亿元)", round(latest["ocf"] / 1e8, 2) if latest.get("ocf") is not None else None),
        ("经营现金流/净利润", latest.get("ocf_to_net_profit")),
        ("商誉占总资产(%)", pct(latest.get("goodwill_ratio"))),
        ("基本每股收益(元)", latest.get("eps")),
    ]
    start = 5
    ws.cell(row=start, column=1, value="最新一期核心指标").font = Font(bold=True)
    for i, (k, v) in enumerate(kv):
        r = start + 1 + i
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)

    # 近 N 年趋势表
    trend_start = start + len(kv) + 3
    ws.cell(row=trend_start, column=1, value="近 %d 年趋势" % len(history)).font = Font(bold=True)
    trend_head = ["报告期", "营收(亿元)", "归母净利(亿元)", "毛利率(%)", "净利率(%)",
                  "ROE(%)", "资产负债率(%)", "经营现金流(亿元)"]
    for c, h in enumerate(trend_head, 1):
        ws.cell(row=trend_start + 1, column=c, value=h)
    _style_header(ws, trend_start + 1, len(trend_head))
    for i, y in enumerate(history):
        r = trend_start + 2 + i
        ws.cell(row=r, column=1, value=y.get("year") or y.get("report_date"))
        ws.cell(row=r, column=2, value=round(y["revenue"] / 1e8, 2) if y.get("revenue") is not None else None)
        ws.cell(row=r, column=3, value=round(y["parent_net_profit"] / 1e8, 2) if y.get("parent_net_profit") is not None else None)
        ws.cell(row=r, column=4, value=pct(y.get("gross_margin")))
        ws.cell(row=r, column=5, value=pct(y.get("net_margin")))
        ws.cell(row=r, column=6, value=pct(y.get("roe")))
        ws.cell(row=r, column=7, value=pct(y.get("debt_ratio")))
        ws.cell(row=r, column=8, value=round(y["ocf"] / 1e8, 2) if y.get("ocf") is not None else None)

    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16


def _write_statement(ws, rows, cn_map):
    """把三大报表数据转置为「科目 × 报告期」表格。"""
    if not rows:
        ws["A1"] = "无数据"
        return
    # 仅保留 cn_map 中确实存在的字段，保持定义顺序（报告期已作为表头，故排除）
    fields = [k for k in cn_map if k != "REPORT_DATE" and any(k in r for r in rows)]
    header = ["科目"] + [str(r.get("REPORT_DATE", "")).split(" ")[0] for r in rows]
    for c, h in enumerate(header, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(header))

    for i, field in enumerate(fields):
        r = i + 2
        ws.cell(row=r, column=1, value=cn_map[field])
        for j, row in enumerate(rows):
            val = row.get(field)
            num = _to_num(val)
            if field in _AMOUNT_FIELDS and isinstance(num, (int, float)):
                num = round(num / 1e8, 2)
            ws.cell(row=r, column=j + 2, value=num)

    ws.column_dimensions["A"].width = 22
    for c in range(2, len(header) + 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 14


def format_statements(data):
    """把三大报表整理成便于前端渲染的结构（中文表头 + 金额换算为亿元）。

    返回 dict：{income|balance|cashflow: {"dates": [...], "fields": [
        {"key", "label", "is_amount", "values": [...]}]}}
    """

    def _fmt(rows, cn_map):
        if not rows:
            return {"dates": [], "fields": []}
        dates = [str(r.get("REPORT_DATE", "")).split(" ")[0] for r in rows]
        fields = []
        for k in cn_map:
            if k == "REPORT_DATE":
                continue
            if not any(k in r for r in rows):
                continue
            values = []
            for r in rows:
                val = _to_num(r.get(k))
                if k in _AMOUNT_FIELDS and isinstance(val, (int, float)):
                    val = round(val / 1e8, 2)
                values.append(val)
            fields.append({
                "key": k,
                "label": cn_map[k],
                "is_amount": k in _AMOUNT_FIELDS,
                "values": values,
            })
        return {"dates": dates, "fields": fields}

    return {
        "income": _fmt(data.get("income") or [], INCOME_CN),
        "balance": _fmt(data.get("balance") or [], BALANCE_CN),
        "cashflow": _fmt(data.get("cashflow") or [], CASHFLOW_CN),
    }


def build_excel(data, indicators):
    """生成 Excel 文件字节流（BytesIO）。"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "核心指标"
    _write_overview(ws1, data, indicators)

    ws2 = wb.create_sheet("利润表")
    _write_statement(ws2, data.get("income") or [], INCOME_CN)

    ws3 = wb.create_sheet("资产负债表")
    _write_statement(ws3, data.get("balance") or [], BALANCE_CN)

    ws4 = wb.create_sheet("现金流量表")
    _write_statement(ws4, data.get("cashflow") or [], CASHFLOW_CN)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
